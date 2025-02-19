import tkinter  as tk
from tkinter import messagebox
from openpyxl import Workbook,load_workbook
import re
import os

saved_book='Info Excel.xlsx'
    #verify if excel file exist
if os.path.exists(saved_book):
    wb = load_workbook(saved_book)
    ws = wb.active
else:
    #Create book of excel
    wb= Workbook()
    ws = wb.active
    ws.append(["Name","Age","Email","Telephone","Address"])


def save_info():
    Name = entry_Name.get()
    Age = entry_Age.get()
    Email = entry_Email.get()
    Telephone = entry_Tel.get()
    Address = entry_Address.get()
    if not Name or not Age or not Email or not Telephone or not Address:
        messagebox.showwarning( title="Advice", message="All the Info must be filled")
        return
    try:
        Age = int(Age)
        Telephone =int(Telephone)
    except ValueError:
        messagebox.showwarning(title="Advice", message="Age and Telephone must be a numbers")
        return
    
    #Email format Checker
    if not re.match(r"[^@]+@[^@]+\.[^@]+", Email):
        messagebox.showwarning(title="Advice", message="Email is not valid")
        return
    ws.append([Name, Age, Email, Telephone, Address])
    wb.save('Info Excel.xlsx')
    messagebox.showinfo(title="Information", message="Information Submit")

    entry_Name.delete(0, tk.END)
    entry_Age.delete(0, tk.END)
    entry_Email.delete(0, tk.END)
    entry_Tel.delete(0, tk.END)
    entry_Address.delete(0, tk.END)


root = tk.Tk()
root.title("Information Input Form")
root.configure(bg='#486587')
label_style = {"bg": '#486587', "fg":"white" }
entry_style= {"bg": '#d3d3d3',"fg":"black"}



label_Name = tk.Label(root, text="Name",**label_style)
label_Name.grid(row=0, column=0, padx=10, pady=5)
entry_Name = tk.Entry(root, **entry_style)
entry_Name.grid(row=0, column=1, padx=10, pady=5)

label_Age = tk.Label(root, text="Age",**label_style)
label_Age.grid(row=1, column=0, padx=10, pady=5)
entry_Age = tk.Entry(root, **entry_style)
entry_Age.grid(row=1, column=1, padx=10, pady=5)

label_Email = tk.Label(root, text="Email",**label_style)
label_Email.grid(row=2, column=0, padx=10, pady=5)
entry_Email = tk.Entry(root, **entry_style)
entry_Email.grid(row=2, column=1, padx=10, pady=5)

label_Tel = tk.Label(root, text="Telephone",**label_style)
label_Tel.grid(row=3, column=0, padx=10, pady=5)
entry_Tel = tk.Entry(root, **entry_style)
entry_Tel.grid(row=3, column=1, padx=10, pady=5)

label_Address = tk.Label(root, text="Address",**label_style)
label_Address.grid(row=4, column=0, padx=10, pady=5)
entry_Address = tk.Entry(root, **entry_style)
entry_Address.grid(row=4, column=1, padx=10, pady=5)

save_button =tk.Button(root, text="Save",command=save_info, bg='#608299', fg='white',width=20)
save_button.grid(row=5,column=0, columnspan=2, padx=10, pady=10)

root.mainloop()